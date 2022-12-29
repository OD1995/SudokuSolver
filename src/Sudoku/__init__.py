from datetime import datetime
from src.SudokuReader.ExcelReader import ExcelReader
from src.Cell import Cell
from math import ceil
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment
from collections import Counter

class Sudoku:

    def __init__(
        self,
        reader_type : str,
        *args,
        **kwargs
    ):
        if reader_type == "excel":
            reader = ExcelReader(*args,**kwargs)
        # else:
        #     print('h2')
        ## Dict
        ##     key: (x,y)
        ##     val: cell value
        self.position_to_value = reader.position_to_value
        ## Dict
        ##     key: int between 1 and 9 inclusive
        ##     val: list of (x,y) coords
        self.value_to_positions = reader.value_to_positions
        self.cells = self.get_cells()

    def get_cells(self):
        cells = {}
        for (x,y),val in self.position_to_value.items():
            cells[(x,y)] = Cell(x,y,val)
        return cells

    def get_new_solveds_by_removing_options_for_surrounding_cells_part1(self,num):
        new_solved = []
        for x,y in self.value_to_positions[num]:
            new_solved.extend(
                self.remove_options_for_surrounding_cells(x,y,num)
            )
        new_solved.extend(self.check_all_groups(num))
        return new_solved

    def get_new_solveds_by_removing_options_for_surrounding_cells_part2(self,new_solved):
        new_solveds_exist = len(new_solved) > 0
        if new_solveds_exist:
            version = 0
            new_solved_dict = {
                version : new_solved
            }
            while new_solveds_exist:
                version += 1
                new_solved_dict[version] = []
                for c in new_solved_dict[version-1]:
                    new_solved_dict[version].extend(
                        self.remove_options_for_surrounding_cells(*c)
                    )
                    new_solved_dict[version].extend(
                        self.check_all_relevant_groups(*c)
                    )
                    self.value_to_positions[c[2]].append((c[0],c[1]))
                new_solveds_exist = len(new_solved_dict[version]) > 0

    def solve(self):
        self.print_sudoku(0,0)
        ## Dict
        ##     key: int between 1 and 9 inclusive
        ##     val: bool, whether all 9 of the key's int has been set
        self.number_fully_solved = {
            num : False
            for num in range(1,10)
        }
        print_count = 0
        last_check = 0
        loop_count = 0
        while not all(self.number_fully_solved.values()):
            for num,fully_solved in self.number_fully_solved.items():
                if not fully_solved:
                    new_solved = self.get_new_solveds_by_removing_options_for_surrounding_cells_part1(num)
                    self.get_new_solveds_by_removing_options_for_surrounding_cells_part2(new_solved)                    
                    self.number_fully_solved[num] = self.is_number_fully_solved(num)
                    # print_count += 1
                    # self.print_sudoku(num,print_count)
            loop_count += 1
            if loop_count % 18 == 0:
                solved_count = self.get_total_solved()
                print(loop_count,solved_count)
                if solved_count > last_check:
                    last_check = solved_count
                else:
                    self.try_narrowing_down()
                    try_something_else1 = self.strategyA()
                    if try_something_else1:
                        try_something_else2 = self.strategyB()
                        if try_something_else2:
                            try_something_else3 = self.strategyC()
                            if try_something_else3:
                                a=1
        self.print_sudoku(0,0,True)

    def strategyC(self):
        ## For any row/column/3x3, if there are two cells which
        ##     have the same two options as its only options, those
        ##     options can be removed from every other cell in the
        ##     row/column/3x3
        success = False
        for p in range(1,10):
            ## Row
            row_coords = [
                (i,p)
                for i in range(1,10)
            ]
            row_res = self.strategyC_inner(row_coords)
            ## Column
            col_coords = [
                (p,i)
                for i in range(1,10)
            ]
            col_res = self.strategyC_inner(col_coords)
            if (row_res or col_res):
                success = True

        for r in range(1,4):
            for c in range(1,4):
                group_coords = self.get_3x3_coords(c,r)
                group_res = self.strategyC_inner(group_coords)
                if group_res:
                    success = True
        return not success

    def strategyC_inner(self,coords_list):
        success = False
        c_cells = [
            self.cells[tuple(z)]
            for z in coords_list
        ]
        two_options = []
        for cc in c_cells:
            if len(cc.options) == 2:
                two_options.append(tuple(cc.options.keys()))
        counter = Counter(two_options)
        mc = counter.most_common()
        if mc[0][1] == 2:
            for cc in c_cells:
                if mc[0][0] != tuple(cc.options.keys()):
                    for o in mc[0][0]:
                        resy = cc.remove_option(o)
                        if self.do_result_check(resy):
                            success = True
        return True

    def strategyB(self):
        ## Work out if in any row/column of 3x3s, the only places 
        ##    a number in two 3x3s can be are on the same column/row,
        ##    Meaning that the third column/row has to be where the
        ##    number is for third 3x3. Or at least it can't be in the
        ##    third 3x3 in the same two column/rows
        key_coords = {}
        for rowo in range(1,4):
            for colo in range(1,4):
                coords = self.get_3x3_coords(colo,rowo)
                non_solved_cells = [
                    self.cells[(x,y)]
                    for x,y in coords
                    if not self.cells[(x,y)].is_solved()
                ]
                for num in range(1,10):
                    xs = []
                    ys = []
                    keyX = f"{rowo},{colo},{num},x"
                    keyY = f"{rowo},{colo},{num},y"
                    for nsv in non_solved_cells:
                        if num in nsv.options:
                            xs.append(nsv.x)
                            ys.append(nsv.y)
                    xs2 = tuple(sorted(list(set(xs))))
                    ys2 = tuple(sorted(list(set(ys))))
                    if len(xs2) == 2:
                        key_coords[keyX] = xs2
                    if len(ys2) == 2:
                        key_coords[keyY] = ys2
        success = False
        ## Columns
        for c in range(1,4):
            for num in range(1,10):
                # if [c,num] == [2,3]:
                #     a=1
                vals = []
                for r in range(1,4):
                    k = f"{r},{c},{num},x"
                    if k in key_coords:
                        vals.append(key_coords[k])
                    else:
                        vals.append(None)
                counter = Counter(vals)
                mc = counter.most_common()[0]
                if (mc[0] is not None) & (mc[1] == 2):
                    ## `num` cannot be in the two columns which `mc` is in
                    ##     and not in the two 3x3s
                    ixix = [
                        i
                        for i,v in enumerate(vals,1)
                        if v != mc[0]
                    ][0]
                    for xx in mc[0]:
                        for yy in self.get_3x3_range(ixix):
                            res = self.cells[(xx,yy)].remove_option(num)
                            if self.do_result_check(res):
                                success = True
        ## Rows
        for r in range(1,4):
            for num in range(1,10):
                # if [c,num] == [2,3]:
                #     a=1
                vals = []
                for c in range(1,4):
                    k = f"{r},{c},{num},y"
                    if k in key_coords:
                        vals.append(key_coords[k])
                    else:
                        vals.append(None)
                counter = Counter(vals)
                mc = counter.most_common()[0]
                if (mc[0] is not None) & (mc[1] == 2):
                    ## `num` cannot be in the two columns which `mc` is in
                    ##     and not in the two 3x3s
                    ixix = [
                        i
                        for i,v in enumerate(vals,1)
                        if v != mc[0]
                    ][0]
                    for yy in mc[0]:
                        for xx in self.get_3x3_range(ixix):
                            res = self.cells[(xx,yy)].remove_option(num)
                            if self.do_result_check(res):
                                success = True
        return not success


    def do_result_check(self,res):
        if res == False:
            return False
        return True


            
    def strategyA(self):
        ## Work out if in any 3x3, all the places a number can be
        ##    are in a row/column. If so, remove that number from the
        ##    options of all the other cells in that row/column, aside
        ##    from the cells in that 3x3
        results = []
        for i in range(1,4):
            for j in range(1,4):
                results.append(self.strategyA_inner(i,j))
        return not any(results)


    def strategyA_inner(self,cix,rix):
        coords = self.get_3x3_coords(cix,rix)
        ## Get empty cells
        # empty_cells = []
        options_dict = {}
        for x,y in coords:
            c = self.cells[(x,y)]
            if not c.is_solved():
                # empty_cells.append(c)
                # all_options.extend(list(c.options.keys()))
                for opt in c.options.keys():
                    if opt in options_dict:
                        options_dict[opt].append((x,y))
                    else:
                        options_dict[opt] = [(x,y)]
        if len(options_dict) == 0:
            return False
        return_me = False
        for uo,coords_list in options_dict.items():
            # if [rix,cix,uo] == [1,2,3]:
            # if [rix,cix,uo] == [3,3,3]:
            #     a=1
            ## Check if all cells that can be `uo` are in the same row/column
            if self.same_row_or_column(coords_list,'row'):
                y = coords_list[0][1]
                for x in range(1,10):
                    if x not in self.get_3x3_range(cix):
                        row_res = self.cells[(x,y)].remove_option(uo)
                        if self.do_result_check(row_res):
                            return_me = True
            if self.same_row_or_column(coords_list,'column'):
                x = coords_list[0][0]
                for y in range(1,10):
                    if y not in self.get_3x3_range(rix):
                        col_res = self.cells[(x,y)].remove_option(uo)
                        if self.do_result_check(col_res):
                            return_me = True
        return return_me

    def same_row_or_column(self,coords_list,axis):
        z = 1 if axis == 'row' else 0
        return all(x[z] == coords_list[0][z] for x in coords_list)

    def try_narrowing_down(
        self
    ):
        groups_of_nine = {}
        for c in self.cells.values():
            for cat in c.get_row_column_3x3():
                if cat not in groups_of_nine:
                    groups_of_nine[cat] = [c]
                else:
                    groups_of_nine[cat].append(c)
        for gn,gon in groups_of_nine.items():
            result = self.narrow_down_options(gn,gon)
            # if result:
            #     a=1

    def narrow_down_options(
        self,
        group_name,
        nine_cells
    ):
        unique_options = {}
        positions = {}
        for cell in nine_cells:
            if cell.value is not None:
                continue

            topt = tuple(cell.options)

            if topt in unique_options:
                unique_options[topt] += 1
            else:
                unique_options[topt] = 1
            
            if topt in positions:
                positions[topt].append((cell.x,cell.y))
            else:
                positions[topt] = [(cell.x,cell.y)]
        max_count = max(unique_options.values())
        if (max_count == 1) or (len(unique_options) == 2):
            return False
        triples = [
            k
            for k,v in unique_options.items()
            if (v == 2) & (len(k) == 3)
        ]
        if len(triples) == 0:
            return False
        triple = triples[0]
        # if len(most_popular) > 3:
        #     a = 1 #will deal with this when it happens
        # elif len(most_popular) == 2:
        #     return False
        # else:
        ## Get all combinations of 2 from the 3 options
        combos = [
            triple[:2],
            triple[-2:],
            (triple[0],triple[2])
        ]
        combo = self.does_combo_exist(combos,unique_options)
        if not combo:
            return False
        pair = [
            triple,
            combo
        ]
        ## Any cell which has options that aren't either `triple` or `combo`
        ##    should have all of the values of `triple` removed from their options
        coords_to_return = []
        for opts,coords in positions.items():
            if opts not in pair:
                coords_to_return.extend(coords)
        
        for val in triple:
            for xy in coords_to_return:
                res = self.cells[xy].remove_option(val)
                # if res:
                #     a=1
        
        return True

    def does_combo_exist(
        self,
        combos,
        unique_options
    ):
        for combo in combos:
            if combo in unique_options:
                if unique_options[combo] != 1:
                    raise ValueError('something has happened')
                return combo
        return False

    def check_all_relevant_groups(
        self,
        x : int,
        y : int,
        num : int
    ):
        more_new_solveds = []
        R = self.check_row(num,y)
        if R:
            more_new_solveds.append(R)
        C = self.check_col(num,x)
        if C:
            more_new_solveds.append(C)
        G = self.check_3x3(
            num,
            ceil(y/3),
            ceil(x/3)
        )
        if G:
            more_new_solveds.append(G)   
        return more_new_solveds     

    def check_all_groups(
        self,
        num : int
    ):
        more_new_solveds = []
        for i in range(1,10):
            R = self.check_row(num,i)
            if R:
                more_new_solveds.append(R)
            C = self.check_col(num,i)
            if C:
                more_new_solveds.append(C)
        for j in range(1,4):
            for k in range(1,4):
                G = self.check_3x3(num,j,k)
                if G:
                    more_new_solveds.append(G)
        return more_new_solveds


    def check_row(
        self,
        value : int,
        row_num : int
    ):
        ## Check if `value` has only one possible answer in the row
        possible_answer = None
        for xc in range(1,10):
            cell = self.cells[xc,row_num]
            if value == cell.value:
                return False
            elif value in cell.options:
                if possible_answer is not None:
                    return False
                else:
                    possible_answer = (xc,row_num)
        ## If we've got to this point then `possible_answer` is the only option
        self.cells[possible_answer[0],possible_answer[1]].set_value(value)
        return (possible_answer[0],possible_answer[1],value)

    def check_col(
        self,
        value : int,
        col_num : int
    ):
        ## Check if `value` has only one possible answer in the col
        possible_answer = None
        for yc in range(1,10):
            cell = self.cells[col_num,yc]
            if value == cell.value:
                return False
            elif value in cell.options:
                if possible_answer is not None:
                    return False
                else:
                    possible_answer = (col_num,yc)
        ## If we've got to this point then `possible_answer` is the only option
        self.cells[possible_answer[0],possible_answer[1]].set_value(value)
        return (possible_answer[0],possible_answer[1],value)

    def check_3x3(
        self,
        value : int,
        row_ix : int,
        col_ix : int
    ):
        if [row_ix,col_ix,value] == [1,3,3]:
            a=1
        ## Check if `value` has only one possible answer in the 3x3
        possible_answer = None
        for x,y in self.get_3x3_coords(col_ix,row_ix):
            cell = self.cells[x,y]
            if value == cell.value:
                return False
            elif value in cell.options:
                if possible_answer is not None:
                    return False
                else:
                    possible_answer = (x,y)
        ## If we've got to this point then `possible_answer` is the only option
        self.cells[possible_answer[0],possible_answer[1]].set_value(value)
        return (possible_answer[0],possible_answer[1],value)


    def get_3x3_coords(
        self,
        col_ix : int,
        row_ix : int
    ):
        return_me = []
        for x in range((3*col_ix)-2,(3*col_ix)+1):
            for y in range((3*row_ix)-2,(3*row_ix)+1):
                return_me.append((x,y))
        return return_me

    def get_3x3_range(self,ix):
        return [x for x in range((3*ix)-2,(3*ix)+1)]

    def remove_options_for_surrounding_cells(
        self,
        x : int,
        y : int,
        num : int
    ):
        # if (x,y) == (4,1):
        #     a=1
        new_solved = []
        ## Remove `num` as an option for all the cells in the same row
        for y0 in range(1,10):
            if y0 != y:
                solved_val = self.cells[(x,y0)].remove_option(num)
                if solved_val:
                    new_solved.append((x,y0,solved_val))
        ## Remove `num` as an option for all the cells in the same row
        for x0 in range(1,10):
            if x0 != x:
                solved_val = self.cells[(x0,y)].remove_option(num)
                if solved_val:
                    new_solved.append((x0,y,solved_val))
        ## Remove `num` as an option for all the cells in the same 3x3
        for x1,y1 in self.cells[(x,y)].get_3x3_coords():
            solved_val = self.cells[(x1,y1)].remove_option(num)
            if solved_val:
                new_solved.append((x1,y1,solved_val))
        return new_solved

    def is_number_fully_solved(
        self,
        num : int
    ):
        return len(self.value_to_positions[num]) == 9

    def to_excel(
        self
    ):
        wb = Workbook()
        ws = wb.active
        filename = f"{datetime.now().timestamp()}.xlsx"
        black = Side(border_style='thin', color='000000')
        white = None#Side(border_style='thin', color='FFFFFF')
        for x in range(1,10):
            for y in range(1,10):
                cell = self.cells[x,y]
                address = f"{get_column_letter(x+1)}{y+1}"
                if cell.value is not None:
                    ws[address] = cell.value
                    ws[address].font = Font(bold=True)
                else:
                    stringValue = ",".join(map(str,cell.options))
                    ws[address] = stringValue
                    ws[address].font = Font(size=5)
                T = black if y == 1 else white
                L = black if x == 1 else white
                B = black if (y % 3) == 0 else white
                R = black if (x % 3) == 0 else white
                ws[address].border = Border(top=T,bottom=B,left=L,right=R)
                ws[address].alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(x+1)].width = 3.429
        wb.save(filename)

    def get_total_solved(self):
        count = 0
        for cell in self.cells.values():
            if cell.is_solved():
                count += 1
        return count

    def print_sudoku(
        self,
        num : int,
        print_count : int,
        A=False
    ):
        Ps = []
        ns = 0
        for y in range(1,10):
            print_me = "|"
            for x in range(1,10):
                raw_val = self.cells[(x,y)].value
                if raw_val is None:
                    val = " "
                else:
                    val = str(raw_val)
                    ns += 1
                print_me += val
                if x % 3 == 0:
                    print_me += "|"
            Ps.append(print_me)
            if y % 3 == 0:
                Ps.append("-"*13)
        print("-------------------------------------")
        print(f"Focusing on {num}")
        print(f"Print count: {print_count}")
        print(f"Numbers solved: {ns}")
        if A:
            print("-"*13)
            for P in Ps:
                print(P)

        if print_count > 50:
            self.try_narrowing_down()
            self.to_excel()
            raise ValueError('too many prints')