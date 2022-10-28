from openpyxl import load_workbook

class ExcelReader:

    def __init__(
        self,
        excel_path : str,
        sheet_name : str
    ) -> None:
        wb = load_workbook(
            filename=excel_path
        )
        ws = wb[sheet_name]
        position_to_value = {}
        value_to_positions = {
            i : []
            for i in range(1,10)
        }
        for x in range(2,11):
            for y in range(2,11):
                val = ws.cell(y,x).value
                X = x - 1
                Y = y - 1
                if val is not None:
                    val = int(val)
                    value_to_positions[val].append((X,Y))
                position_to_value[(X,Y)] = val
        self.position_to_value = position_to_value
        self.value_to_positions = value_to_positions